import sys
import os
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def main():
    print("Starting Time Counting...")

def load_json(file_path):
    if os.path.exists(file_path):
        with open(file_path, 'r') as file:
            data = json.load(file)
        print("File loaded successfully.")
    else:
        initial_data = {
            "IsLoggedOn": False,
            "Days": []
        }
        with open(file_path, 'w') as file:
            json.dump(initial_data, file)
        print(f"The file {file_path} was not found, so it was created.")
        data = initial_data
    return data

def save_json(file_path, data):
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)
    print(f"JSON data saved successfully at: {os.path.abspath(file_path)}")

def logon_user(data):
    if data['IsLoggedOn']:
        print("User is already logged on.")
    else:
        current_time = datetime.now().strftime('%H%M')
        current_date = datetime.now().strftime('%d.%m.%Y')
        log_entry = {
            "day": current_date,
            "logon-time": current_time,
            "logout-time": "",
            "time-worked": "",
            "logout-reason": ""
        }
        data['Days'].append(log_entry)
        data['IsLoggedOn'] = True
        print(f"User logged on at {current_time} on {current_date}.")

def logoff_user(data, reason="hours full"):
    if not data['IsLoggedOn']:
        print("User is not currently logged on.")
    else:
        current_time = datetime.now().strftime('%H%M')
        last_entry = data['Days'][-1]
        if last_entry['logout-time']:
            print("User has already logged off.")
            return

        last_entry['logout-time'] = current_time
        logon_time = datetime.strptime(last_entry['logon-time'], '%H%M')
        logout_time = datetime.strptime(current_time, '%H%M')
        time_worked = logout_time - logon_time
        last_entry['time-worked'] = f"{time_worked.seconds // 3600:02}:{(time_worked.seconds // 60) % 60:02}"
        last_entry['logout-reason'] = reason
        data['IsLoggedOn'] = False
        print(f"User logged off at {current_time}. Time worked: {last_entry['time-worked']}. Reason: {reason}.")


def get_times(data, date=None, month=None, full_export=False):
    if date:
        entries = [day for day in data['Days'] if day['day'] == date]
        if entries:
            print(f"Entries for {date}:")
            print("\n")
            for entry in entries:
                print(entry)
            print("\n")
        else:
            print(f"No entries found for {date}.")
    elif month:
        month_entries = [day for day in data['Days'] if datetime.strptime(day['day'], '%d.%m.%Y').strftime('%B').lower() == month.lower() or datetime.strptime(day['day'], '%d.%m.%Y').strftime('%m') == month.zfill(2)]
        if month_entries:
            last_day = None
            print(f"Entries for {month.capitalize()}:")
            print("\n")
            for entry in month_entries:
                if entry['day'] != last_day:
                    if last_day is not None:
                        print("-" * 80)
                    last_day = entry['day']
                print(entry)
            print("\n")
            print("-" * 80)
            print("\n")
        else:
            print(f"No entries found for {month}.")
    elif full_export:
        excel_file = "workinghours_export.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Working Hours"
        
        headers = ['Day', 'Logon Time', 'Logout Time', 'Time Worked', 'Logout Reason']
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}1"]
            cell.value = header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, day in enumerate(data['Days'], 2):
            sheet[f"A{row_num}"].value = day['day']
            sheet[f"B{row_num}"].value = day['logon-time']
            sheet[f"C{row_num}"].value = day['logout-time']
            sheet[f"D{row_num}"].value = day['time-worked']
            sheet[f"E{row_num}"].value = day['logout-reason']
        
        workbook.save(excel_file)
        print(f"Excel data exported successfully at: {os.path.abspath(excel_file)}")

if len(sys.argv) > 1:
    command = sys.argv[1]
    file_path = 'workinghours.json'
    data = load_json(file_path)
    
    if command == "-logon":
        logon_user(data)
    elif command == "-logoff":
        reason = sys.argv[2] if len(sys.argv) > 2 else "hours full"
        logoff_user(data, reason)
    elif command == "-get-times":
        if len(sys.argv) > 2:
            sub_command = sys.argv[2]
            if sub_command == "Date":
                if len(sys.argv) > 3:
                    date = sys.argv[3]
                    get_times(data, date=date)
                else:
                    print("Date argument missing.")
            elif sub_command == "Month":
                if len(sys.argv) > 3:
                    month = sys.argv[3]
                    get_times(data, month=month)
                else:
                    print("Month argument missing.")
            elif sub_command == "Full-Export":
                get_times(data, full_export=True)
            else:
                print("Invalid argument for get-times.")
        else:
            print("No argument provided for get-times.")
    elif command == "-help":
        box_width = 80
        border = "+" + "-" * (box_width - 2) + "+"
        
        title = "Available Commands:"
        padding = (box_width - len(title) - 2) // 2  
        
        print(border)
        print("|" + " " * padding + title + " " * (box_width - len(title) - padding - 2) + "|")
        print(border)

        commands = [
            "-logon               =>      Logs your start time",
            "-logoff              =>      Logs you off with the argument 'hours full'",
            " -> LOGOFF-REASON    =>      Adds a reason for the Logoff to the logoff",
            "-get-times           =>      A Command for Exporting the Times",
            " -> Date DD.MM.YYYY  =>      Exports the time of a specific day",
            " -> month (MM)       =>      Exports the times of that specific month",
            " -> Full-Export      =>      Exports all entries as an xlsx file",
            "-help                =>      Shows all commands"
        ]

        for command in commands:
            print(f"| {command.ljust(box_width - 4)} |")

        print(border)
    else:
        print("Unknown command.")
    
    save_json(file_path, data)
else:
    print("No command provided.")